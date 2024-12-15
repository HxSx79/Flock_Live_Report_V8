"""Module for recording line crossings to Excel"""
from datetime import datetime
import openpyxl
import os
from typing import List
from .line_detector import LineCrossing
from .bom_reader import BOMReader

class CrossingRecorder:
    def __init__(self, filename: str = "flock_report.xlsx"):
        self.filename = filename
        self.bom_reader = BOMReader()
        self._ensure_file_exists()
        
    def _ensure_file_exists(self):
        """Create Excel file with headers if it doesn't exist"""
        if not os.path.exists(self.filename):
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = ["Class Name", "Program", "Part Number", "Part Description", "Day", "Time"]
            ws.append(headers)
            wb.save(self.filename)
    
    def record_crossings(self, crossings: List[LineCrossing]) -> None:
        """Record multiple line crossings to Excel"""
        if not crossings:
            return
            
        try:
            wb = openpyxl.load_workbook(self.filename)
            ws = wb.active
            
            for crossing in crossings:
                # Get part information
                part_info = self.bom_reader.get_part_info(crossing.class_name)
                
                # Get current timestamp
                now = datetime.now()
                
                # Prepare row data
                row_data = [
                    crossing.class_name,
                    part_info['program'],
                    part_info['part_number'],
                    part_info['description'],
                    now.strftime("%Y-%m-%d"),
                    now.strftime("%H:%M:%S")
                ]
                
                # Insert new row after header
                ws.insert_rows(2)
                
                # Write data
                for col, value in enumerate(row_data, start=1):
                    ws.cell(row=2, column=col, value=value)
            
            # Save all changes at once
            wb.save(self.filename)
            
        except Exception as e:
            print(f"Error recording crossings: {e}")
            # Re-create file if corrupted
            if "Invalid file" in str(e):
                self._ensure_file_exists()